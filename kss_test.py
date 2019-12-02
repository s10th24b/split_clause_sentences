import os
import glob
import platform
import kss
import openpyxl
import time
import re

dir_splt = '/'
clause_col = ''
end_of_row = ''
ws = ''

def init_env():
    print("Initializing Environment . . .")
    process_time = time.time()
    global dir_splt
    if platform.system() == 'Windows':
        dir_splt = '\\'
    else:
        dir_splt = '/'
    print("Current System:",platform.system())
    print("directory split:",dir_splt)
    currentdir = os.getcwd()
    parentdir = os.path.dirname(currentdir)
    datadir = currentdir+dir_splt+'data'
    outputdir = currentdir+dir_splt+'output_data'
    files = glob.glob(datadir+dir_splt+'*.xlsx')
    #filepath = datadir+dir_splt+'test_data.xlsx'
    #xlsx_file = openpyxl.load_workbook(filepath,data_only=True)
    #xlsx_file = openpyxl.load_workbook(filepath,data_only=True)
    
    if not (os.path.isdir(datadir)):
        os.makedirs(os.path.join(datadir))
    if not (os.path.isdir(outputdir)):
        os.makedirs(os.path.join(outputdir))
    process_time = time.time() - process_time
    print("Initializing Completed! in %.3f" %(process_time))
    return dir_splt,currentdir,parentdir,datadir,outputdir,files


def load_xlsx(f):
    current_xlsx = openpyxl.load_workbook(f,data_only=True)
    return current_xlsx

def get_clause_cols(xlsx_file):
    print("Processing get_clause_cols . . .")
    process_time = time.time()
    global clause_col
    global end_of_row
    global ws

    ws = xlsx_file['Sheet1'] #ws means work_sheet

    clause_col= 'E'
    end_of_row = len(ws[clause_col])
    print("end_of_row:",end_of_row)

    clause_cols = ws[clause_col+str(2):clause_col+str(end_of_row)] #from start to end
    process_time = time.time() - process_time
    print("get_clause_cols Completed! in %.3f" %(process_time))
    return clause_cols

def fix_splt_sent(s_list):
    for idx,s in enumerate(s_list):
        #print("s:",s)
        regex = re.compile("^\w+\.$")
        mc = regex.findall(s)
        if len(mc) >0:
            print("kss Exception Occured! Matched:",mc)
            print("Original Whole Clause:",s_list)
            try:
                s_list[idx+1] = s_list[idx]+' '+s_list[idx+1]
            except IndexError:
                print("Index",idx+1,"not exists.")
            del s_list[idx]
            print("Final s_list:",s_list)


def splt_sents(cells):
    print("Processing splt_sents . . .")
    process_time = time.time()
    global ws
    current_row = 2
    for row in cells:
        for cell in row:
            if current_row % 100 == 0:
                print("current_row:",current_row)
            s = cell.value
            #print("s:",s)
            s_list = kss.split_sentences(s)
            fix_splt_sent(s_list)





            sent_len = len(s_list)
            if sent_len == 0:
                print("current_row:",current_row)
                print("sent_len is zero. Is it empty?")
                current_row+=1
            elif sent_len == 1:
                #print("current_row:",current_row)
                #print("sent_len is one. normal")
                current_row+=1
            elif sent_len > 1: #sent_len is bigger than 1
                print("Gotcha! sent_len: ",sent_len,"in",current_row)
                print("Original Whole Clause:",s_list)
                for idx,sent in enumerate(s_list):
                    seq = idx + 1
                    ws[clause_col+str(current_row)] = sent
                    if sent_len - seq > 0:
                        #print("insert_rows")
                        ws.insert_rows(current_row+1)
                    current_row+=1
            else:
                print("current_row:",current_row)
                print("Abnormal")
    process_time = time.time() - process_time
    print("get_clause_cols Completed! in %.3f" %(process_time))

if __name__ == "__main__":
    main_process_time = time.time()
    dir_splt,currentdir,parentdir,datadir,outputdir,files = init_env()
    for idx,f in enumerate(files):
        filename = str(f.split(dir_splt)[len(f.split(dir_splt))-1])
        #print("filename:",filename)
        print("\n============================================================\n")
        print("Processing File:",filename,". . .")

        current_file = load_xlsx(f)
        cl_cols = get_clause_cols(current_file)
        splt_sents(cl_cols)
        current_file.save(outputdir+dir_splt+filename+'_output.xlsx')
        print("processing",filename,"is completed!")
        print("\n============================================================\n")
    main_process_time = time.time() - main_process_time
    print("Program Ended! in %.3f" %(main_process_time))
